"""
传感器BOM 导入脚本
从 Excel 文件导入 BOM 数据到数据库

Excel 文件结构：
- Sheet 名: Sheet2
- 列: 物料代码, 物料版本, 物料名称, 需求数量
- 最后一行是产品自身行（物料代码 == 文件名）
- 其他行是该产品的零部件

映射规则：
- BOM 表:
    id              = product_code (如 "4472000005")
    product_code    = product_code
    product_name    = product_name (取自最后一行物料名称)
    version         = product_version (取自最后一行物料版本)
    status          = "active"
    created_at      = 当前时间
- BOMItem 表:
    bom_id          = product_code
    material_code   = 物料代码
    material_name   = 物料名称
    specification   = 物料版本
    quantity        = 需求数量 (float)
    material_type   = "component"
"""
import os
import glob
from datetime import datetime

import openpyxl
from sqlalchemy.exc import IntegrityError

from src.models.database import get_db_session, init_db
from src.models.tables import BOM, BOMItem


def parse_quantity(value):
    """将 Excel 中的数量字段转为浮点数，无法转换则返回 0"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (ValueError, TypeError):
        return 0.0


def import_sensor_bom(file_path: str, db) -> dict:
    """导入单个传感器BOM Excel 文件"""
    file_name = os.path.basename(file_path)
    product_code = os.path.splitext(file_name)[0]

    wb = openpyxl.load_workbook(file_path, data_only=True)
    if "Sheet2" not in wb.sheetnames:
        return {"file": file_name, "success": False, "error": "未找到 Sheet2"}
    ws = wb["Sheet2"]

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"file": file_name, "success": False, "error": "空文件"}

    # 最后一行是产品自身
    header_row = rows[-1]
    product_version = str(header_row[1]).strip() if header_row[1] is not None else ""
    product_name = str(header_row[2]).strip() if header_row[2] is not None else ""

    # 中间行为零部件（跳过表头与最后一行）
    item_rows = rows[1:-1]

    # 检查产品 BOM 是否已存在 -> 存在则删除旧 items 与 header 后重建
    existing_header = (
        db.query(BOM).filter(BOM.product_code == product_code).first()
    )
    if existing_header:
        db.query(BOMItem).filter(BOMItem.bom_id == product_code).delete()
        db.delete(existing_header)
        db.flush()

    # 创建 BOM header
    header = BOM(
        id=product_code,
        product_code=product_code,
        product_name=product_name,
        version=product_version,
        status="active",
        created_at=datetime.utcnow(),
    )
    db.add(header)
    db.flush()

    # 创建 BOM items
    item_count = 0
    for row in item_rows:
        if not row or row[0] is None:
            continue
        material_code = str(row[0]).strip()
        if not material_code:
            continue
        material_version = str(row[1]).strip() if row[1] is not None else ""
        material_name = str(row[2]).strip() if row[2] is not None else ""

        item = BOMItem(
            bom_id=product_code,
            material_code=material_code,
            material_name=material_name,
            specification=material_version,
            quantity=parse_quantity(row[3]),
            material_type="component",
        )
        db.add(item)
        item_count += 1

    return {
        "file": file_name,
        "success": True,
        "product_code": product_code,
        "product_name": product_name,
        "product_version": product_version,
        "item_count": item_count,
    }


def main():
    workspace = os.path.dirname(os.path.abspath(__file__))
    excel_files = sorted(glob.glob(os.path.join(workspace, "44*.xlsx")))

    if not excel_files:
        print("未找到任何 Excel 文件")
        return

    print(f"发现 {len(excel_files)} 个 Excel 文件待导入")
    print("-" * 60)

    init_db()
    db = get_db_session()
    try:
        results = []
        for f in excel_files:
            try:
                r = import_sensor_bom(f, db)
                results.append(r)
            except IntegrityError as e:
                db.rollback()
                results.append(
                    {"file": os.path.basename(f), "success": False, "error": str(e)}
                )
            except Exception as e:
                db.rollback()
                results.append(
                    {"file": os.path.basename(f), "success": False, "error": str(e)}
                )

        db.commit()

        print("\n导入结果：")
        print("-" * 60)
        success_count = 0
        total_items = 0
        for r in results:
            if r.get("success"):
                success_count += 1
                total_items += r["item_count"]
                print(
                    f"[OK] {r['file']}  产品={r['product_code']} ({r['product_name']}) "
                    f"版本={r['product_version']}  零部件数={r['item_count']}"
                )
            else:
                print(f"[FAIL] {r['file']}  错误={r.get('error')}")
        print("-" * 60)
        print(f"合计：成功 {success_count}/{len(results)}，总零部件 {total_items} 条")
    finally:
        db.close()


if __name__ == "__main__":
    main()